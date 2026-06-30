from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
import math
import re
import uuid
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

TOLERANCE = 0.01
GS_FLAG_TEXT = "GS involvement indicated"
EXCEPTION_HEADERS = [
    "Trade ID", "Broker", "119 Row", "Mgmt C Row", "119 Brokerage", "Mgmt C Brokerage",
    "Difference 119 - Mgmt C", "Sum of Fees", "Fees", "Discount %", "GS Involvement Flag",
    "Status", "Comparison Direction",
]

BROKER_MAP = {
    "AS400": "ARR", "AKH": "ADAM", "ALT": "STEVEN", "AMM": "Alexandre", "ANT": "ANTHONY",
    "ASC": "ANDY", "AVA": "Jack Avallone", "AXL": "Alex Lacroix", "BAU": "David Bauer",
    "BES": "Bridgeview Desk", "BHN": "BRANDON", "BIL": "BILL", "BRU": "ALBERT",
    "BRY": "BRADY", "CFO": "CHRISTOPHER", "DAF": "DAVE", "DAJ": "Darren",
    "DAS": "Dan Simonetti", "DBU": "David Bauer", "DEI": "GEORGE", "DLG": "Dan Lago",
    "DOH": "DOUG", "ERZ": "ERIC", "EVS": "IAN", "GIL": "JAMES", "GRG": "Mike Grgich",
    "GTM": "MAURICE", "JCP": "JAKE", "JDM": "JOE", "JKP": "John Perrotta BL",
    "JMC": "JORDAN", "JOE": "JLEG", "JOS": "STOLARZ", "JOW": "JOSH", "KIM": "KIM",
    "KOW": "DANKO", "KUV": "KURT", "LAN": "Robby", "LAX": "Harrison Lax",
    "LEP": "Andrew Lepore", "LGO": "DAN", "LYN": "JACK", "MAC": "Thomas McManus Swaps",
    "MCM": "Tom McManus Funding", "MDG": "MICHAEL", "MFA": "Michael Ferrara", "MGT": "Pierrick",
    "MIR": "REGAZZI", "MKN": "Matthew", "MLY": "MICHAEL", "MSK": "Mickey Sacks",
    "MVR": "MICHAEL V", "NEX": "Nexus Citi Singles", "NIL": "ANDERS", "NVA": "NOVA",
    "OWL": "OWEN", "PHL": "Philip", "PIE": "Pierre", "RBJ": "Robert Baccala",
    "RBK": "Richard", "RDA": "REDA ABBADI", "RHC": "Herman Rockefeller", "RSR": "RORY",
    "SKM": "TERENCE", "SNI": "JOES", "TGG": "Tim Gregor", "TGR": "TIM", "THO": "THOMAS",
    "TOD": "Todd", "TRS": "ROUSE", "VGD": "ALEX", "WHJ": "JAMES", "WIG": "GHEEN",
    "ZAA": "ZACH", "ZUL": "Jason Zulin",
}

@dataclass
class ReconciliationResult:
    result_id: str
    output_path: str
    filename: str
    latest_date_label: str
    reporting_period: str
    tabs: Dict[str, List[Dict[str, Any]]]
    metrics: Dict[str, Any]
    summary: Dict[str, Any]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s_.\-\r\n]+", "", str(value or "").strip().lower())


def to_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    s = str(value).strip().replace(",", "").replace("$", "").replace("£", "").replace("€", "")
    s = s.replace("%", "").replace("(", "-").replace(")", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def clean_trade_id(value: Any) -> str:
    s = "" if value is None else str(value).strip()
    p = s.find("-")
    if p > 0:
        s = s[:p]
    if len(s) > 10 and s[:10].isdigit():
        s = s[:10]
    try:
        f = float(s)
        if f.is_integer():
            s = str(int(f))
        else:
            s = str(f).split(".", 1)[0]
    except ValueError:
        pass
    return clean_text(s)


def col_index(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def header_map(ws: Worksheet, row: int = 1) -> Dict[str, int]:
    result: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = normalize_header(ws.cell(row, c).value)
        if key and key not in result:
            result[key] = c
    return result


def find_header(ws: Worksheet, header: str, row: int = 1) -> int:
    return header_map(ws, row).get(normalize_header(header), 0)


def find_main_sheet(wb: Workbook, required: Iterable[str], ignore: Iterable[str]) -> Optional[Worksheet]:
    ignored = {x.lower() for x in ignore}
    for ws in wb.worksheets:
        if ws.title.lower() in ignored:
            continue
        if all(find_header(ws, h) for h in required):
            return ws
    return None


def missing_headers(ws: Worksheet, required: Iterable[str]) -> List[str]:
    return [h for h in required if not find_header(ws, h)]


def safe_cell(ws: Worksheet, row: int, col: int) -> Any:
    if col < 1 or col > ws.max_column:
        return None
    return ws.cell(row, col).value


def ordinal(n: int) -> str:
    if 10 <= n % 100 <= 20:
        return "th"
    return {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")


def latest_trade_date(ws: Worksheet) -> Optional[date]:
    # VBA uses original 119 fixed columns: AC Year Trade, AD Month, AE Day.
    year_col, month_col, day_col = col_index("AC"), col_index("AD"), col_index("AE")
    latest: Optional[date] = None
    for r in range(2, ws.max_row + 1):
        y, m, d = safe_cell(ws, r, year_col), safe_cell(ws, r, month_col), safe_cell(ws, r, day_col)
        try:
            candidate = date(int(float(y)), int(float(m)), int(float(d)))
        except Exception:
            continue
        if latest is None or candidate > latest:
            latest = candidate
    return latest


def latest_date_label_from_filename(filename: str) -> str:
    m = re.search(r"Thru\s+(.+?)(?:\.|$)", filename, flags=re.IGNORECASE)
    return m.group(1).strip() if m else ""


def latest_date_label(filename: str, ws119: Worksheet) -> str:
    from_name = latest_date_label_from_filename(filename)
    if from_name:
        return from_name
    d = latest_trade_date(ws119)
    if d:
        return f"{d.day}{ordinal(d.day)} {d.strftime('%b %Y')}"
    today = date.today()
    return f"{today.day}{ordinal(today.day)} {today.strftime('%b %Y')}"


def reporting_period(ws119: Worksheet) -> str:
    d = latest_trade_date(ws119)
    if not d:
        return ""
    return f"Reporting Period: 1st {d.strftime('%b')} to {d.day}{ordinal(d.day)} {d.strftime('%b %Y')}"



def first_row_header_map(ws: Worksheet) -> Dict[str, int]:
    try:
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return {}
    result: Dict[str, int] = {}
    for idx, value in enumerate(row, start=1):
        key = normalize_header(value)
        if key and key not in result:
            result[key] = idx
    return result


def find_main_sheet_stream(wb: Workbook, required: Iterable[str], ignore: Iterable[str]) -> Optional[Worksheet]:
    ignored = {x.lower() for x in ignore}
    for ws in wb.worksheets:
        if ws.title.lower() in ignored:
            continue
        h = first_row_header_map(ws)
        if all(normalize_header(req) in h for req in required):
            return ws
    return None


def row_value(row: Tuple[Any, ...], col: int) -> Any:
    idx = col - 1
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def latest_date_label_from_date(d: Optional[date]) -> str:
    if d:
        return f"{d.day}{ordinal(d.day)} {d.strftime('%b %Y')}"
    today = date.today()
    return f"{today.day}{ordinal(today.day)} {today.strftime('%b %Y')}"


def reporting_period_from_date(d: Optional[date]) -> str:
    if not d:
        return ""
    return f"Reporting Period: 1st {d.strftime('%b')} to {d.day}{ordinal(d.day)} {d.strftime('%b %Y')}"


def process_119(path: Path) -> Tuple[List[Dict[str, Any]], str, Optional[date]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = find_main_sheet_stream(wb, ["Deal Nb", "Desk"], ["BrokerList", "Pivot", "Adjusted_119_InMemory"])
    if ws is None:
        raise ValueError("119 file must contain a main sheet with headers including Deal Nb and Desk.")

    h = first_row_header_map(ws)
    def hcol(name: str) -> int:
        return h.get(normalize_header(name), 0)

    base_missing = [x for x in ["Deal Nb", "Desk", "Net in LC", "Discount", "Fees", "Amount W", "Amount X", "Amount Z"] if not hcol(x)]
    net_bro_col = hcol("Net Bro") or hcol("Net Bro.")
    if not net_bro_col:
        base_missing.append("Net Bro")
    if base_missing:
        raise ValueError("119 file is missing required column(s): " + ", ".join(base_missing))

    net_lc_col = hcol("Net in LC")
    discount_col = hcol("Discount")
    fees_col = hcol("Fees")
    amount_w_col = hcol("Amount W")
    amount_x_col = hcol("Amount X")
    amount_z_col = hcol("Amount Z")
    prod_col = col_index("C")
    dealer_col = col_index("H")
    account_col = col_index("BU")
    year_col, month_col, day_col = col_index("AC"), col_index("AD"), col_index("AE")

    latest: Optional[date] = None
    agg: Dict[str, Dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            candidate = date(int(float(row_value(row, year_col))), int(float(row_value(row, month_col))), int(float(row_value(row, day_col))))
            if latest is None or candidate > latest:
                latest = candidate
        except Exception:
            pass

        account = row_value(row, account_col)
        trade_id = clean_text("2" + ("" if account is None else str(account).strip()))
        dealer_code = clean_text(row_value(row, dealer_col))
        broker = BROKER_MAP.get(dealer_code, "")
        if not trade_id or not broker:
            continue

        net_bro = to_number(row_value(row, net_bro_col))
        amount_w = to_number(row_value(row, amount_w_col))
        amount_x = to_number(row_value(row, amount_x_col))
        amount_z = to_number(row_value(row, amount_z_col))
        prod = clean_text(row_value(row, prod_col))
        bro = net_bro + amount_z + (amount_x if prod in {"EQEXT", "EQEXI"} else amount_w)

        key = f"{broker}|{trade_id}"
        if key not in agg:
            agg[key] = {
                "Broker": broker, "ID": trade_id, "Sum of Net in LC": 0.0, "Sum of Discount": 0.0,
                "Sum of Fees": 0.0, "Sum of Net Bro": 0.0, "Sum of Amount W": 0.0,
                "Sum of Amount X": 0.0, "Sum of Amount Z": 0.0, "Sum of Bro": 0.0,
            }
        out = agg[key]
        out["Sum of Net in LC"] += to_number(row_value(row, net_lc_col))
        out["Sum of Discount"] += to_number(row_value(row, discount_col))
        out["Sum of Fees"] += to_number(row_value(row, fees_col))
        out["Sum of Net Bro"] += net_bro
        out["Sum of Amount W"] += amount_w
        out["Sum of Amount X"] += amount_x
        out["Sum of Amount Z"] += amount_z
        out["Sum of Bro"] += bro

    adjusted = [v for v in agg.values() if v["Sum of Bro"] < -0.5 or v["Sum of Bro"] > 0.5]
    wb.close()
    return adjusted, ws.title, latest


def process_mgmt(path: Path) -> Tuple[List[Dict[str, Any]], str]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = find_main_sheet_stream(wb, ["Trade ID", "Brokerage"], ["TRS", "Pivot", "Brokerage Data", "Adjusted_MgmtC_InMemory"])
    if ws is None:
        raise ValueError("Mgmt C file must contain a main sheet with headers including Trade ID and Brokerage.")
    h = first_row_header_map(ws)
    trade_id_col = h.get(normalize_header("Trade ID"), 0)
    brokerage_col = h.get(normalize_header("Brokerage"), 0)
    missing = []
    if not trade_id_col:
        missing.append("Trade ID")
    if not brokerage_col:
        missing.append("Brokerage")
    if missing:
        raise ValueError("Mgmt C file is missing required column(s): " + ", ".join(missing))

    source_broker_col = col_index("D")
    agg: Dict[str, Dict[str, Any]] = {}
    prior_broker = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_broker = str(row_value(row, source_broker_col) or "").strip()
        if raw_broker:
            prior_broker = raw_broker
        broker = clean_text(prior_broker)
        trade_id = clean_trade_id(row_value(row, trade_id_col))
        if not trade_id:
            continue
        key = f"{broker}|{trade_id}"
        if key not in agg:
            agg[key] = {"Broker": broker, "Trade ID": trade_id, "Total": 0.0}
        agg[key]["Total"] += to_number(row_value(row, brokerage_col))

    adjusted = [v for v in agg.values() if v["Total"] < -0.5 or v["Total"] > 0.5]
    adjusted.sort(key=lambda x: x["Trade ID"])
    wb.close()
    return adjusted, ws.title

def build_key(trade_id: Any, broker: Any) -> str:
    return f"{clean_text(trade_id)}|{clean_text(broker)}"


def discount_pct(discount: Any, net_lc: Any) -> float:
    d, n = to_number(discount), to_number(net_lc)
    return 0.0 if d == 0 or n == 0 else d / n


def gs_flag(pct: float) -> str:
    return GS_FLAG_TEXT if 0.03 <= abs(pct) <= 0.10 else ""


def fees_value(diff: Any, sum_fees: Any) -> Any:
    if diff == "" or diff is None:
        return ""
    d = to_number(diff)
    return d - to_number(sum_fees) if abs(d) > 0.1 else 0.0


def exception_row(trade_id: Any, broker: Any, row119: Any, row_mgmt: Any, brk119: Any, brk_mgmt: Any,
                  diff: Any, sum_fees: Any, pct: Any, flag: Any, status: str, direction: str) -> Dict[str, Any]:
    return {
        "Trade ID": clean_text(trade_id),
        "Broker": clean_text(broker),
        "119 Row": row119,
        "Mgmt C Row": row_mgmt,
        "119 Brokerage": brk119,
        "Mgmt C Brokerage": brk_mgmt,
        "Difference 119 - Mgmt C": diff,
        "Sum of Fees": sum_fees,
        "Fees": fees_value(diff, sum_fees),
        "Discount %": pct,
        "GS Involvement Flag": flag,
        "Status": status,
        "Comparison Direction": direction,
    }


def reconcile(adjusted_119: List[Dict[str, Any]], adjusted_mgmt: List[Dict[str, Any]]) -> Tuple[Dict[str, List[Dict[str, Any]]], Dict[str, Any]]:
    dict119: Dict[str, float] = {}
    dict119_rows: Dict[str, int] = {}
    dict119_pct: Dict[str, float] = {}
    dict119_flag: Dict[str, str] = {}
    dict119_fees: Dict[str, float] = {}
    dict_mgmt: Dict[str, float] = {}
    dict_mgmt_rows: Dict[str, int] = {}
    id119, id_mgmt = set(), set()
    dup119, dup_mgmt = set(), set()
    metrics = {
        "Matched trades with no exception": 0,
        "Brokerage differences found on 119 vs Mgmt C": 0,
        "Brokerage differences found on Mgmt C vs 119": 0,
        "Trades in 119 source report missing from Mgmt C source report": 0,
        "Trades in Mgmt C source report missing from 119 source report": 0,
        "Broker mismatches from 119 side": 0,
        "Broker mismatches from Mgmt C side": 0,
        "Duplicate 119 Trade ID + Broker keys": 0,
        "Duplicate Mgmt C Trade ID + Broker keys": 0,
        "GS involvement flags detected": 0,
        "Rows where Fees formula was triggered": 0,
        "Total absolute brokerage difference": 0.0,
        "Estimated total Fees adjustment from 119 vs Mgmt C": 0.0,
    }

    for idx, row in enumerate(adjusted_119, start=2):
        trade_id, broker = clean_text(row.get("ID")), clean_text(row.get("Broker"))
        if trade_id:
            id119.add(trade_id)
        if trade_id and broker:
            key = build_key(trade_id, broker)
            pct = discount_pct(row.get("Sum of Discount"), row.get("Sum of Net in LC"))
            flag = gs_flag(pct)
            if flag:
                metrics["GS involvement flags detected"] += 1
            if key in dict119:
                dup119.add(key)
                metrics["Duplicate 119 Trade ID + Broker keys"] += 1
            else:
                dict119[key] = to_number(row.get("Sum of Bro"))
                dict119_rows[key] = idx
                dict119_pct[key] = pct
                dict119_flag[key] = flag
                dict119_fees[key] = to_number(row.get("Sum of Fees"))

    for idx, row in enumerate(adjusted_mgmt, start=2):
        trade_id, broker = clean_text(row.get("Trade ID")), clean_text(row.get("Broker"))
        if trade_id:
            id_mgmt.add(trade_id)
        if trade_id and broker:
            key = build_key(trade_id, broker)
            if key in dict_mgmt:
                dup_mgmt.add(key)
                metrics["Duplicate Mgmt C Trade ID + Broker keys"] += 1
            else:
                dict_mgmt[key] = to_number(row.get("Total"))
                dict_mgmt_rows[key] = idx

    rows119: List[Dict[str, Any]] = []
    rows_mgmt: List[Dict[str, Any]] = []
    all_rows: List[Dict[str, Any]] = []

    for idx, row in enumerate(adjusted_119, start=2):
        trade_id, broker = clean_text(row.get("ID")), clean_text(row.get("Broker"))
        if not trade_id:
            continue
        key = build_key(trade_id, broker)
        brk119 = to_number(row.get("Sum of Bro"))
        sum_fees = to_number(row.get("Sum of Fees"))
        pct = discount_pct(row.get("Sum of Discount"), row.get("Sum of Net in LC"))
        flag = gs_flag(pct)
        if not broker:
            status, brk_mgmt, diff, row_m = "Missing Broker in 119", "", "", ""
        elif key in dup119:
            status, brk_mgmt, diff, row_m = "Duplicate Trade ID + Broker in 119", "", "", ""
        elif key in dict_mgmt:
            brk_mgmt = dict_mgmt[key]
            diff = brk119 - brk_mgmt
            if abs(diff) <= TOLERANCE and not flag:
                metrics["Matched trades with no exception"] += 1
                continue
            status = "Brokerage Difference" if abs(diff) > TOLERANCE else "Discount % Flag - GS Involvement"
            row_m = dict_mgmt_rows[key]
            if abs(diff) > TOLERANCE:
                metrics["Brokerage differences found on 119 vs Mgmt C"] += 1
                metrics["Total absolute brokerage difference"] += abs(diff)
            if abs(diff) > 0.1:
                metrics["Rows where Fees formula was triggered"] += 1
                metrics["Estimated total Fees adjustment from 119 vs Mgmt C"] += diff - sum_fees
        elif trade_id in id_mgmt:
            status, brk_mgmt, diff, row_m = "Trade ID matched, Broker mismatch or missing from Mgmt C", "", "", ""
            metrics["Broker mismatches from 119 side"] += 1
        else:
            status, brk_mgmt, diff, row_m = "Trade ID missing from Mgmt C", "", "", ""
            metrics["Trades in 119 source report missing from Mgmt C source report"] += 1
        ex = exception_row(trade_id, broker, idx, row_m, brk119, brk_mgmt, diff, sum_fees, pct, flag, status, "119 vs Mgmt C")
        rows119.append(ex); all_rows.append(ex)

    for idx, row in enumerate(adjusted_mgmt, start=2):
        trade_id, broker = clean_text(row.get("Trade ID")), clean_text(row.get("Broker"))
        if not trade_id:
            continue
        key = build_key(trade_id, broker)
        brk_mgmt = to_number(row.get("Total"))
        if not broker:
            status, brk119, diff, sum_fees, pct, flag, row119 = "Missing Broker in Mgmt C", "", "", "", "", "", ""
        elif key in dup_mgmt:
            status, brk119, diff, sum_fees, pct, flag, row119 = "Duplicate Trade ID + Broker in Mgmt C", "", "", "", "", "", ""
        elif key in dict119:
            brk119 = dict119[key]
            diff = brk119 - brk_mgmt
            pct, flag, sum_fees, row119 = dict119_pct[key], dict119_flag[key], dict119_fees[key], dict119_rows[key]
            if abs(diff) <= TOLERANCE and not flag:
                continue
            status = "Brokerage Difference" if abs(diff) > TOLERANCE else "Discount % Flag - GS Involvement"
            if abs(diff) > TOLERANCE:
                metrics["Brokerage differences found on Mgmt C vs 119"] += 1
        elif trade_id in id119:
            status, brk119, diff, sum_fees, pct, flag, row119 = "Trade ID matched, Broker mismatch or missing from 119", "", "", "", "", "", ""
            metrics["Broker mismatches from Mgmt C side"] += 1
        else:
            status, brk119, diff, sum_fees, pct, flag, row119 = "Trade ID missing from 119", "", "", "", "", "", ""
            metrics["Trades in Mgmt C source report missing from 119 source report"] += 1
        ex = exception_row(trade_id, broker, row119, idx, brk119, brk_mgmt, diff, sum_fees, pct, flag, status, "Mgmt C vs 119")
        rows_mgmt.append(ex); all_rows.append(ex)

    gs_rows = [r for r in all_rows if r.get("GS Involvement Flag") == GS_FLAG_TEXT]
    issues = [r.copy() for r in all_rows if not r.get("GS Involvement Flag")]
    metrics["GS Discount rows created from All Exceptions"] = len(gs_rows)
    metrics["Issues rows after initial blank GS Involvement Flag selection"] = len(issues)

    seen = set()
    deduped = []
    for r in issues:
        if r["Status"] == "Brokerage Difference":
            k = build_key(r["Trade ID"], r["Broker"])
            if k in seen:
                continue
            seen.add(k)
        deduped.append(r)
    issues = deduped

    def keep_diff(r: Dict[str, Any]) -> bool:
        diff = r.get("Difference 119 - Mgmt C")
        if diff == "" or diff is None:
            return True
        return not (-1 <= to_number(diff) <= 1)
    issues = [r for r in issues if keep_diff(r)]
    metrics["Issues rows after Difference 119 - Mgmt C filter"] = len(issues)

    def keep_fee(r: Dict[str, Any]) -> bool:
        diff = r.get("Difference 119 - Mgmt C")
        if diff == "" or diff is None:
            return True
        fee = r.get("Fees")
        return not (fee != "" and fee is not None and -1 <= to_number(fee) <= 1)
    issues = [r for r in issues if keep_fee(r)]

    def final_keep(r: Dict[str, Any]) -> bool:
        if "missing from 119" in str(r.get("Status", "")).lower():
            return True
        pct = r.get("Discount %")
        if pct == "" or pct is None:
            return False
        return abs(to_number(pct)) < 0.0000001
    issues = [r for r in issues if final_keep(r)]
    for r in issues:
        # Final Issues column G must always equal column E minus column F.
        # Treat a blank 119/Mgmt C brokerage as zero when the opposite side exists,
        # matching how reviewers expect the final Excel Issues tab to behave.
        brk119, brk_mgmt = r.get("119 Brokerage"), r.get("Mgmt C Brokerage")
        if brk119 not in ("", None) or brk_mgmt not in ("", None):
            diff = to_number(brk119) - to_number(brk_mgmt)
            r["Difference 119 - Mgmt C"] = diff
            r["Fees"] = fees_value(diff, r.get("Sum of Fees"))
        r["Comments"] = ""
    metrics["Final Issues rows after Fees filter and Discount % = 0% rule"] = len(issues)

    metrics["Total absolute brokerage difference"] = round(metrics["Total absolute brokerage difference"], 2)
    metrics["Estimated total Fees adjustment from 119 vs Mgmt C"] = round(metrics["Estimated total Fees adjustment from 119 vs Mgmt C"], 2)

    return {
        "119 vs Mgmt C": rows119,
        "Mgmt C vs 119": rows_mgmt,
        "All Exceptions": all_rows,
        "GS Discount": gs_rows,
        "Issues": issues,
    }, metrics


def add_rows(ws: Worksheet, headers: List[str], rows: List[Dict[str, Any]], start_row: int = 1) -> None:
    for c, h in enumerate(headers, start=1):
        ws.cell(start_row, c, h)
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c, h in enumerate(headers, start=1):
            ws.cell(r_idx, c, row.get(h, ""))


def style_sheet(ws: Worksheet, header_row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(vertical="top")
    ws.freeze_panes = f"A{header_row + 1}"
    sample_max = min(ws.max_row, header_row + 200)
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for r in range(1, sample_max + 1):
            max_len = max(max_len, len(str(ws.cell(r, col).value or "")))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 42)
    for col_name in ["E", "F", "G", "H", "I"]:
        for row in range(header_row + 1, ws.max_row + 1):
            ws[f"{col_name}{row}"].number_format = '#,##0.00'
    for row in range(header_row + 1, ws.max_row + 1):
        ws[f"J{row}"].number_format = '0.00%'

def build_workbook(output_path: Path, tabs: Dict[str, List[Dict[str, Any]]], metrics: Dict[str, Any], summary: Dict[str, Any], reporting_period_label: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "119 vs Mgmt C Month End Reconciliation Summary"
    ws["A1"].font = Font(bold=True, size=16)
    summary_rows = [
        ("119 source workbook", summary["119 source workbook"]),
        ("Mgmt C source workbook", summary["Mgmt C source workbook"]),
        ("Run date/time", summary["Run date/time"]),
        ("Primary matching rule", "Match Trade ID first, then Broker."),
        ("Brokerage comparison rule", "Difference 119 - Mgmt C = 119 Brokerage - Mgmt C Brokerage."),
        ("Discount % rule", "IF(Sum of Discount = 0 or Sum of Net in LC = 0, 0, Sum of Discount / Sum of Net in LC)."),
        ("Fees rule", "IF(ABS(Difference 119 - Mgmt C) > 0.1, Difference 119 - Mgmt C - Sum of Fees, 0)."),
        ("GS Involvement Flag", "Populated when absolute Discount % is between 3% and 10% inclusive."),
    ]
    row = 3
    for k, v in summary_rows:
        ws.cell(row, 1, k); ws.cell(row, 2, v); row += 1
    row += 1
    ws.cell(row, 1, "Metric"); ws.cell(row, 2, "Value"); row += 1
    for k, v in metrics.items():
        ws.cell(row, 1, k); ws.cell(row, 2, v); row += 1
    row += 2
    ws.cell(row, 1, "Output Tab"); ws.cell(row, 2, "Explanation"); row += 1
    explanations = {
        "Summary": "Explains source files, business rules, and reconciliation metrics.",
        "119 vs Mgmt C": "Exceptions found while looping through adjusted 119 rows.",
        "Mgmt C vs 119": "Exceptions found while looping through adjusted Mgmt C rows.",
        "All Exceptions": "Combined exceptions from both comparison directions.",
        "GS Discount": "Rows from All Exceptions where the GS involvement flag is populated.",
        "Issues": "Final reviewer issues after blank GS flag, duplicate, difference, fees, and discount filters.",
        "Approval": "Professional sign-off page for approver name, date, and comments.",
    }
    for k, v in explanations.items():
        ws.cell(row, 1, k); ws.cell(row, 2, v); row += 1
    style_sheet(ws, 13)

    for tab_name in ["119 vs Mgmt C", "Mgmt C vs 119", "All Exceptions", "GS Discount"]:
        ws = wb.create_sheet(tab_name)
        add_rows(ws, EXCEPTION_HEADERS, tabs.get(tab_name, []))
        style_sheet(ws)

    ws = wb.create_sheet("Issues")
    ws["A1"] = reporting_period_label or "Reporting Period:"
    ws["A1"].font = Font(bold=True)
    issue_headers = EXCEPTION_HEADERS + ["Comments"]
    add_rows(ws, issue_headers, tabs.get("Issues", []), start_row=2)
    # Ensure Excel column G is visibly driven by columns E and F on the Issues tab.
    for row_num in range(3, ws.max_row + 1):
        ws[f"G{row_num}"] = f'=IF(AND(E{row_num}="",F{row_num}=""),"",N(E{row_num})-N(F{row_num}))'
        ws[f"G{row_num}"].number_format = '#,##0.00'
    style_sheet(ws, header_row=2)
    # Hide internal columns like the VBA macro, while still leaving them available.
    for col in ["C", "D", "H", "K"]:
        ws.column_dimensions[col].hidden = True

    ws = wb.create_sheet("Approval")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "Reconciliation Sign-Off"
    ws["A1"].font = Font(bold=True, size=22, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="17385E")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A3:F3")
    ws["A3"] = "This reconciliation has been reviewed and approved by:"
    ws["A3"].font = Font(italic=True, color="38506B", size=12)
    ws["A3"].alignment = Alignment(horizontal="center")

    field_fill = PatternFill("solid", fgColor="EEF5FC")
    border_side = Side(style="thin", color="B7C9DC")
    box_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    fields = [("A5", "Approver Name", "B5:F5"), ("A7", "Date Signed", "B7:F7"), ("A9", "Comments", "B9:F14")]
    for label_cell, label, merge_range in fields:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, color="17385E")
        ws[label_cell].alignment = Alignment(vertical="top")
        ws.merge_cells(merge_range)
        start = merge_range.split(":")[0]
        ws[start].fill = field_fill
        ws[start].border = box_border
        ws[start].alignment = Alignment(wrap_text=True, vertical="top")

    ws["A16"] = "Reviewer Checklist"
    ws["A16"].font = Font(bold=True, color="17385E", size=13)
    checklist = [
        "Summary metrics reviewed",
        "Issues tab reviewed and comments completed",
        "GS Discount tab reviewed",
        "Final workbook approved for month-end evidence",
    ]
    for idx, item in enumerate(checklist, start=17):
        ws[f"A{idx}"] = "☐"
        ws[f"B{idx}"] = item
        ws[f"B{idx}"].font = Font(color="38506B")

    ws.column_dimensions["A"].width = 22
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 18
    wb.save(output_path)


def preview_tabs(tabs: Dict[str, List[Dict[str, Any]]], limit: int = 500) -> Dict[str, List[Dict[str, Any]]]:
    return {k: v[:limit] for k, v in tabs.items()}


def run_reconciliation(file119_path: Path, mgmt_path: Path, file119_name: str, mgmt_name: str, output_dir: Path) -> ReconciliationResult:
    adjusted_119, sheet119_name, latest_dt = process_119(file119_path)
    adjusted_mgmt, sheet_mgmt_name = process_mgmt(mgmt_path)
    tabs, metrics = reconcile(adjusted_119, adjusted_mgmt)
    label = latest_date_label_from_filename(file119_name) or latest_date_label_from_date(latest_dt)
    reporting = reporting_period_from_date(latest_dt)
    run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary = {
        "119 source workbook": file119_name,
        "Mgmt C source workbook": mgmt_name,
        "119 main sheet": sheet119_name,
        "Mgmt C main sheet": sheet_mgmt_name,
        "Run date/time": run_time,
        "Adjusted 119 rows": len(adjusted_119),
        "Adjusted Mgmt C rows": len(adjusted_mgmt),
    }
    result_id = uuid.uuid4().hex
    filename = f"119 vs Mgmt C Thru {label}.xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{result_id}_{filename}"
    full_tabs = {"Summary": [], **tabs, "Approval": []}
    build_workbook(output_path, tabs, metrics, summary, reporting)
    return ReconciliationResult(
        result_id=result_id,
        output_path=str(output_path),
        filename=filename,
        latest_date_label=label,
        reporting_period=reporting,
        tabs=preview_tabs(full_tabs),
        metrics=metrics,
        summary=summary,
    )
